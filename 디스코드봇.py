import discord
import asyncio
import random 
import openpyxl
import urllib
import urllib.request
import bs4
import datetime
import random
client = discord.Client ()
token = 'NTQwNDE3NDMyNDU2NzkwMDM2.DzReeg.vZIvPMr60rctTDCpejMhoPgyyec'
@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("------------")
    await client.change_presence(game=discord.Game(name='디스코드봇', type=1))

@client.event
async def on_message(message):
    if message.author.bot: #만약 메시지를 보낸사람이 봇일 경우에는
        return None #동작하지 않고 무시합니다.

    id = message.author.id #id라는 변수에는 메시지를 보낸사람의 ID를 담습니다.
    channel = message.channel #channel이라는 변수에는 메시지를 받은 채널의 ID를 담습니다.

@client.event
async def on_message(message): #잠만여네
     
    id = message.author.id
    channel = message.channel
    owner = ['417123204469882890','361091925266137089']
    if message.author.bot:
          return None
     
    if message.content.startswith('!안녕'):
        await client.send_message(channel,'안녕하세요')

    if message.content.startswith("!사다리"):
        team = message.content[5:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname)
        for i in range(0, len(person)):
            await client.send_message(message.channel, person[i] + "------>" + teamname[i])

    if message.content.startswith("!학습"):
         file = openpyxl.load_workbook("기억.xlsx")
         sheet = file.active
         learn = message.content.split(" ")
         for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
              sheet["A" + str(i)].value = learn[1]
              sheet["B" + str(i)].value = learn[2]
              break
         await client.send_message(message.channel, "기억완료!")
         file.save("기억.xlsx")
    if message.content.startswith("!기억") and not message.content.startswith("!기억삭제"):
         file = openpyxl.load_workbook("기억.xlsx")
         sheet = file.active
         memory = message.content.split(" ")
         for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith("!기억삭제"):
         file = openpyxl.load_workbook("기억.xlsx")
         sheet = file.active
         memory = message.content.split(" ")
         for i in range(1, 51):
             if sheet["A" + str(i)].value == str(memory[1]):
                 sheet["A" + str(i)].value = "-"
                 sheet["B" + str(i)].value = "-"
                 await client.send_message(message.channel, "기억삭제완료!")
                 file.save("기억.xlsx")
                 break

    if "흥" in message.content or "ㅡㅡ" in message.content or  "바보" in message.content: 

        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 51):
            if str(sheet["A"+ str(i)].value) == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value)  == "-":
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 1
                break

        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고를 받았습니다. 단어선택에 주의해주세요.")

    if message.content.startswith("!경고"):
        memid = message.content.split(" ")
        file = openpyxl.load_workbook("경고.xlsx")
        member = discord.utils.get(client.get_all_members(),id=memid[1])
        sheet = file.active
        for i in range(1, 51):
            if str(sheet["A"+ str(i)].value) == str(memid[1]):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                if int(sheet["B" + str(i)].value) == 3:
                    await client.ban(member, 1)
                break
            if str(sheet["A" + str(i)].value)  == "-":
                sheet["A" + str(i)].value = str(memid[1])
                sheet["B" + str(i)].value = 1
                break

        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고를 받았습니다. 단어선택에 주의해주세요.")

    if message.content.startswith("!배그솔로"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        solo1 = bsObj.find("div", {"class": "overview"})
        solo2 = solo1.text
        solo3 = solo2.strip()
        channel = message.channel
        embed = discord.Embed(
            title='배그솔로 정보',
            description='배그솔로 정보입니다.',
            colour=discord.Colour.green())
        if solo3 == "No record":
            print("솔로 경기가 없습니다.")
            embed.add_field(name='배그를 한판이라도 해주세요', value='솔로 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            solo4 = solo1.find("span", {"class": "value"})
            soloratting = solo4.text  # -------솔로레이팅---------
            solorank0_1 = solo1.find("div", {"class": "grade-info"})
            solorank0_2 = solorank0_1.text
            solorank = solorank0_2.strip()  # -------랭크(그마,브론즈)---------

            print("레이팅 : " + soloratting)
            print("등급 : " + solorank)
            print("")
            embed.add_field(name='레이팅', value=soloratting, inline=False)
            embed.add_field(name='등급', value=solorank, inline=False)

            soloKD1 = bsObj.find("div", {"class": "kd stats-item stats-top-graph"})
            soloKD2 = soloKD1.find("p", {"class": "value"})
            soloKD3 = soloKD2.text
            soloKD = soloKD3.strip()  # -------킬뎃(2.0---------
            soloSky1 = soloKD1.find("span", {"class": "top"})
            soloSky2 = soloSky1.text  # -------상위10.24%---------

            print("킬뎃 : " + soloKD)
            print("킬뎃상위 : " + soloSky2)
            print("")
            embed.add_field(name='킬뎃,킬뎃상위', value=soloKD + " " + soloSky2, inline=False)
            # embed.add_field(name='킬뎃상위', value=soloSky2, inline=False)

            soloWinRat1 = bsObj.find("div", {"class": "stats"})  # 박스
            soloWinRat2 = soloWinRat1.find("div", {"class": "winratio stats-item stats-top-graph"})
            soloWinRat3 = soloWinRat2.find("p", {"class": "value"})
            soloWinRat = soloWinRat3.text.strip()  # -------승률---------
            soloWinRatSky1 = soloWinRat2.find("span", {"class": "top"})
            soloWinRatSky = soloWinRatSky1.text.strip()  # -------상위?%---------

            print("승률 : " + soloWinRat)
            print("승률상위 : " + soloWinRatSky)
            print("")
            embed.add_field(name='승률,승률상위', value=soloWinRat + " " + soloWinRatSky, inline=False)
            # embed.add_field(name='승률상위', value=soloWinRatSky, inline=False)

            soloHead1 = soloWinRat1.find("div", {"class": "headshots stats-item stats-top-graph"})
            soloHead2 = soloHead1.find("p", {"class": "value"})
            soloHead = soloHead2.text.strip()  # -------헤드샷---------
            soloHeadSky1 = soloHead1.find("span", {"class": "top"})
            soloHeadSky = soloHeadSky1.text.strip()  # # -------상위?%---------

            print("헤드샷 : " + soloHead)
            print("헤드샷상위 : " + soloHeadSky)
            print("")
            embed.add_field(name='헤드샷,헤드샷상위', value=soloHead + " " + soloHeadSky, inline=False)
            # embed.add_field(name='헤드샷상위', value=soloHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("!배그듀오"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "duo modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그듀오 정보',
            description='배그듀오 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('듀오 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='듀오 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)

            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD + " " + duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    if message.content.startswith("!배그스쿼드"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "squad modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그스쿼드 정보',
            description='배그스쿼드 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('스쿼드 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='스쿼드 경기 전적이 없습니다..', inline=False)
            await client.send_message(channel, embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)

            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD + " " + duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await client.send_message(channel, embed=embed)

    # 투표설정
    if message.content.startswith('!투표설정'):
        order = message.content[6]

        file_parti = openpyxl.load_workbook("참여자.xlsx")
        file_vote = openpyxl.load_workbook("투표목록.xlsx")
        sheet_parti = file_parti.active
        sheet_vote = file_vote.active

        flag = True
        flag_preEnd = False
        if sheet_parti["A1"].value != 0:
            flag_preEnd = True
        if order == '0':
            sheet_parti["A1"] = 1
        elif order == '1':
            sheet_parti["A1"] = 2
        else:
            await client.send_message(channel, '잘못된 입력 형식입니다.')
            flag = False

        if flag:
            # 이전에 진행중이었던 투표가 있었다면 결과 출력
            if flag_preEnd:
                await client.send_message(channel, '이전의 투표를 종료합니다.')
                flag = True
                if sheet_parti["A1"].value == 0:
                    await client.send_message(channel, '종료할 투표가 존재하지 않습니다.')
                    flag = False

                if flag:
                    vote = []

                    n = sheet_vote["A1"].value

                    for i in range(2, 2 + n):
                        vote.append([sheet_vote["B" + str(i)].value, sheet_vote["A" + str(i)].value])

                    vote.sort(reverse=True)

                    des = ''
                    tmp1 = 0
                    tmp2 = 0
                    for i in range(0, n):
                        if tmp1 != vote[i][0]:
                            tmp2 = i + 1
                            tmp1 = vote[i][0]
                        des = des + "%d" % tmp2 + '위 - <' + vote[i][1] + '> --------------- ' + "%d" % vote[i][
                            0] + "표" + '\n' + '\n'

                    embed = discord.Embed(title="이전투표결과", description=des, color=0xff0000)
                    embed.set_footer(text='투표 총 참여자 - ' + "%d" % sheet_parti["B1"].value + "명")
                    await client.send_message(channel, embed=embed)

            # 참여자 목록 초기화
            sheet_parti["B1"] = 0
            for i in range(2, 32):
                sheet_parti["A" + str(i)].value = '-'
                sheet_parti["B" + str(i)].value = 0
            file_parti.save("참여자.xlsx")

            tmp = message.content[8:]
            candidate = tmp.split('/')

            # 투표 목록 초기화
            sheet_vote["A1"] = len(candidate)

            cnt = 2
            for i in candidate:
                sheet_vote["A" + str(cnt)].value = i
                sheet_vote["B" + str(cnt)].value = 0
                cnt += 1

            file_vote.save("투표목록.xlsx")

            # 투표 목록 출력
            des = ''
            cnt = 1
            for i in candidate:
                des = des + "%d" % cnt + '번째 후보 : ' + i + '\n' + '\n'
                cnt += 1
            embed = discord.Embed(title="후보", description=des, color=0x00ff00)
            embed.set_image(
                url="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT-4pFRHcvahEWCQkH-horx1XFydmx3w3ZtNn2wwfOu5KB5IxfY")
            embed.set_footer(text='우덜식 민주주의')

            await client.send_message(channel, embed=embed)



    # 투표참여
    elif message.content.startswith('!투표참여'):
        file = openpyxl.load_workbook("참여자.xlsx")
        file_vote = openpyxl.load_workbook("투표목록.xlsx")
        file_private = openpyxl.load_workbook("비공개투표.xlsx")

        sheet = file.active
        sheet_vote = file_vote.active
        sheet_private = file_private.active

        per = sheet["B1"].value + 1

        if sheet["A1"].value == 0:
            await client.send_message(channel, '참여할 투표가 존재하지 않습니다.')
        else:
            cnt_member = 1
            flag = True
            for i in range(2, 2 + per):
                if str(sheet["A" + str(i)].value) == '-':
                    sheet["A" + str(i)].value = str(id)
                    break
                elif str(sheet["A" + str(i)].value) == str(id):
                    await client.send_message(channel, '이미 투표에 참여되었습니다')
                    flag = False
                    break

            if flag:
                sheet["B1"].value += 1
                n_parti = sheet["B1"].value
                n_vote = sheet_vote["A1"].value
                file.save("참여자.xlsx")
                await client.send_message(channel, '<@' + id + '>님이 투표에 참여하셨습니다.')

                if sheet["A1"].value == 2:
                    t = random.randint(0, n_vote)
                    for i in range(1, n_vote + 1):
                        sheet_private.cell(n_parti, i).value = ((i + t) % n_vote) + 1
                    file_private.save("비공개투표.xlsx")
                    await client.send_message(channel, '암호화 된 번호를 보냈습니다.')
                    me = await client.get_user_info(id)

                    des = '암호화 된 번호입니다.\n\n'
                    for i in range(1, n_vote + 1):
                        des = des + '후보 <' + str(sheet_vote["A" + str(i + 1)].value) + '> ---------- ' + str(
                            sheet_private.cell(n_parti, i).value) + '번\n\n'
                    embed = discord.Embed(title="암호화 된 번호 목록", description=des, color=0x00ffff)
                    await client.send_message(me, embed=embed)



    # 투표종료
    elif message.content.startswith('!투표종료'):
        file = openpyxl.load_workbook("투표목록.xlsx")
        sheet = file.active

        file_parti = openpyxl.load_workbook("참여자.xlsx")
        sheet_parti = file_parti.active

        flag = True
        if sheet_parti["A1"].value == 0:
            await client.send_message(channel, '종료할 투표가 존재하지 않습니다.')
            flag = False

        if flag:
            vote = []

            n = sheet["A1"].value

            for i in range(2, 2 + n):
                vote.append([sheet["B" + str(i)].value, sheet["A" + str(i)].value])

            vote.sort(reverse=True)

            des = ''
            tmp1 = 0
            tmp2 = 0
            for i in range(0, n):
                if tmp1 != vote[i][0]:
                    tmp2 = i + 1
                    tmp1 = vote[i][0]
                des = des + "%d" % tmp2 + '위 - <' + vote[i][1] + '> --------------- ' + "%d" % vote[i][
                    0] + "표" + '\n' + '\n'

            embed = discord.Embed(title="투표결과", description=des, color=0xff0000)
            embed.set_footer(text='투표 총 참여자 - ' + "%d" % sheet_parti["B1"].value + "명")
            await client.send_message(channel, embed=embed)

            # 참여자 목록에 투표종료를 저장
            sheet_parti["A1"].value = 0
            file_parti.save("참여자.xlsx")


    # 투표
    elif message.content.startswith('!투표'):
        file_parti = openpyxl.load_workbook("참여자.xlsx")
        file_vote = openpyxl.load_workbook("투표목록.xlsx")
        file_private = openpyxl.load_workbook("비공개투표.xlsx")

        sheet_parti = file_parti.active
        sheet_vote = file_vote.active
        sheet_private = file_private.active

        flag1 = True
        if sheet_parti["A1"].value == 0:
            await client.send_message(channel, '참여할 투표가 존재하지 않습니다.')
            flag1 = False
        elif message.content[3] != ' ':
            await client.send_message(channel, '잘못된 입력 형식입니다.')
            flag1 = False

        if flag1:
            n = sheet_vote["A1"].value
            per = sheet_parti["B1"].value
            idx = int(message.content[4:])  # 사용자가 투표한 번호

            flag = False
            t = 0
            idx_parti = 0
            for i in range(2, 2 + per):
                if str(sheet_parti["A" + str(i)].value) == str(id):
                    flag = True
                    t = sheet_parti["B" + str(i)].value
                    idx_parti = i
                    break

            if flag:
                if idx <= 0 or idx > n:
                    await client.send_message(channel, '없는 후보입니다.')
                else:
                    if t != 0:
                        sheet_vote["B" + str(t + 1)].value -= 1

                    if sheet_parti["A1"].value == 2:
                        for i in range(1, n + 1):
                            if sheet_private.cell(idx_parti - 1, i).value == idx:
                                idx = i
                                break

                    sheet_vote["B" + str(idx + 1)].value += 1
                    sheet_parti["B" + str(idx_parti)].value = idx

                    file_parti.save("참여자.xlsx")
                    file_vote.save("투표목록.xlsx")
                    await client.send_message(channel, '투표 완료')
            else:
                await client.send_message(channel, '투표에 참여하지 않았습니다. 투표에 참여 후 다시 시도해주세요.')
client.run(token)



